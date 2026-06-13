"""
Cones Section OEE ETL Pipeline
================================
Reads daily Cones OEE Excel files and produces a single JSON file
that the Cones OEE dashboard (cones_oee_dashboard.html) reads directly.

REQUIREMENTS
------------
Install once, in Command Prompt:
    py -m pip install openpyxl pandas

USAGE
-----
Place all your daily Excel files in one folder, then run:

    py oee_etl.py

By default it looks for Excel files in the same folder as this script
and writes oee_data.json next to it.

To specify a different folder:
    py oee_etl.py --folder "D:\WORKING\CONES_OEE\\CONES_data\\OEE Files"

To write the JSON somewhere else:
    py oee_etl.py --output "D:\WORKING\CONES_OEE\\output\\oee_data.json"

FILE NAMING
-----------
Each Excel file must contain the date somewhere in its name.
Accepted formats:  DD.MM.YYYY  DD_MM_YYYY  DD-MM-YYYY  DD.MM.YY
Examples that all work:
    14.05.26.xlsx
    14.05.2026.xlsx
    cones_14_05_2026.xlsx

FORMULA METHODOLOGY  (3-factor OEE, no cavity term)
--------------------
  OEE_running  = sum(actual_good) / sum(target)           [aggregate ratio]
  Availability = sum(run_hrs) / sum(shift_hrs)            [aggregate ratio]
  Quality      = sum(good) / sum(good + broken_gen_pcs)   [broken_gen only, not broken_used]
  Performance  = OEE_running / (Availability x Quality)   [back-calculated residual]
                 => guarantees Avail x Perf x Quality = OEE_running exactly
                 => negative perf_loss means machines ran above rated speed
  Shift run %  = sum(hours_run) / (total_slots x 12)
  Overall OEE  = OEE_running x Shift_run_pct
"""

import os
import re
import sys
import json
import argparse
import logging
from datetime import datetime
from pathlib import Path

# ── Dependency check ──────────────────────────────────────────────────────────
missing = []
try:
    import pandas as pd
except ImportError:
    missing.append("pandas")
try:
    from openpyxl import load_workbook
except ImportError:
    missing.append("openpyxl")

if missing:
    print("\nERROR: Missing required libraries. Please run this in Command Prompt:")
    print(f"\n    python -m pip install {' '.join(missing)}\n")
    sys.exit(1)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("oee_etl")

# ── Column positions in OEE tracker (1-indexed, for openpyxl) ────────────────
CONES_COL = {
    "date":            2,   # B
    "shift":           3,   # C
    "duration_hrs":    6,   # F
    "machine_no":      7,   # G
    "operator_name":   8,   # H
    "to_fill":         9,   # I
    "product_type":    10,  # J
    "cone_size":       11,  # K
    "cone_type":       12,  # L
    "weight_gm":       14,  # N
    "rated_speed":     15,  # O
    "target_pcs":      16,  # P
    "actual_good_pcs": 17,  # Q
    "broken_gen_kg":   18,  # R
    "broken_used_kg":  19,  # S
    "total_rej_pcs":   20,  # T
    "run_hrs":         24,  # X
    "dt_pump_clean":   25,  # Y
    "dt_mould_chg":    26,  # Z
    "dt_variety_co":   27,  # AA  <- variety changeover (new)
    "dt_gas_filter":   28,  # AB  <- gas filter cleaning (new)
    "dt_plate_clean":  29,  # AC
    "dt_proc_fail":    30,  # AD
    "dt_mach_bkdn":    31,  # AE
    "dt_manpower":     32,  # AF
    "dt_rm_pm":        33,  # AG
    "dt_power_cut":    34,  # AH  <- power cut (new)
    "dt_other":        35,  # AI
}

SHIFT_COL = {
    "date":      2,  # B
    "shift":     3,  # C
    "machine":   4,  # D
    "ran":       5,  # E
    "reason":    6,  # F
    "hours_run": 7,  # G  (Y=12 hardcoded, N=0 hardcoded, P=partial hours entered)
}

DT_KEYS   = ["dt_pump_clean", "dt_mould_chg", "dt_variety_co", "dt_gas_filter",
             "dt_plate_clean", "dt_proc_fail", "dt_mach_bkdn", "dt_manpower",
             "dt_rm_pm", "dt_power_cut", "dt_other"]
DT_LABELS = ["Pump cleaning", "Mould change", "Variety changeover", "Gas filter cleaning",
             "Plate cleaning", "Process failure", "Mach. breakdown", "Manpower unavail.",
             "RM/PM shortage", "Power cut", "Other"]
DT_PLANNED = {"Pump cleaning", "Mould change", "Variety changeover",
              "Gas filter cleaning", "Plate cleaning"}


# ── Date extraction ───────────────────────────────────────────────────────────

def extract_date(path):
    """
    Find a date in the filename and return as datetime.
    Accepts: DD.MM.YYYY  DD_MM_YYYY  DD-MM-YYYY  DD.MM.YY
    """
    stem = Path(path).stem
    # Try 4-digit year first
    m = re.search(r"(\d{1,2})[._\-](\d{1,2})[._\-](\d{4})", stem)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # Try 2-digit year
    m = re.search(r"(\d{1,2})[._\-](\d{1,2})[._\-](\d{2})$", stem)
    if m:
        try:
            year = 2000 + int(m.group(3))
            return datetime(year, int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    log.warning(f"Could not parse date from: {Path(path).name}")
    return None


# ── Excel parsing ─────────────────────────────────────────────────────────────

def parse_file(path):
    """
    Parse one daily Cones OEE Excel file.
    Returns dict with keys: date_str, all_slots, not_run, records
    or None if the file cannot be parsed.
    """
    file_date = extract_date(path)
    if file_date is None:
        return None

    date_str = file_date.strftime("%Y-%m-%d")
    log.info(f"Reading: {Path(path).name}  ->  {date_str}")

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        log.error(f"  Cannot open file: {e}")
        return None

    if "Shift tracker" not in wb.sheetnames or "OEE tracker" not in wb.sheetnames:
        log.error(f"  Missing required sheets. Found: {wb.sheetnames}")
        return None

    # ── Shift tracker ──────────────────────────────────────────────────────
    ws_shift = wb["Shift tracker"]
    not_run  = []
    all_slots = []

    for row in range(15, 300):
        machine = ws_shift.cell(row=row, column=SHIFT_COL["machine"]).value
        shift   = ws_shift.cell(row=row, column=SHIFT_COL["shift"]).value
        ran     = ws_shift.cell(row=row, column=SHIFT_COL["ran"]).value

        if machine is None or shift is None:
            if row > 20:
                break
            continue

        try:
            machine_no = int(machine)
        except (ValueError, TypeError):
            continue

        shift_str  = str(shift).strip()
        ran_str    = str(ran).strip().upper() if ran is not None else "Y"
        reason_raw = ws_shift.cell(row=row, column=SHIFT_COL["reason"]).value
        reason_str = str(reason_raw).strip() if reason_raw else None

        # Y = full shift (12 hrs hardcoded), N = 0 hrs hardcoded, P = read col G
        if ran_str == "Y":
            hours_run = 12.0
        elif ran_str == "N":
            hours_run = 0.0
        else:  # P — partial: read col G
            hrs_raw = ws_shift.cell(row=row, column=SHIFT_COL["hours_run"]).value
            try:
                hours_run = float(hrs_raw) if hrs_raw is not None else 0.0
            except (ValueError, TypeError):
                hours_run = 0.0

        slot = {
            "date":       date_str,
            "shift":      shift_str,
            "machine_no": machine_no,
            "hours_run":  hours_run,
            "ran_flag":   ran_str,
            # Capture reason for N and P slots; blank for Y (fully ran)
            "reason":     (reason_str or "Missing reason code") if ran_str in ("N", "P") else "",
        }
        all_slots.append(slot)

        if ran_str == "N":
            not_run.append({
                "date":       date_str,
                "shift":      shift_str,
                "machine_no": machine_no,
                "reason":     reason_str or "Missing reason code",
            })

    log.info(f"  Shift tracker: {len(all_slots)} slots, {len(not_run)} not run")

    # ── OEE tracker ────────────────────────────────────────────────────────
    ws_oee  = wb["OEE tracker"]
    records = []

    # Find last row with data
    last_row = 11
    for _r in range(11, 2000):
        has_data = any(
            ws_oee.cell(row=_r, column=c).value is not None
            for c in [2, 7, 16, 17]  # date, machine, target, actual
        )
        if has_data:
            last_row = _r

    for row in range(11, last_row + 1):
        def get(col_name, row=row):
            return ws_oee.cell(row=row, column=CONES_COL[col_name]).value

        def num(col_name, default=0, row=row):
            v = ws_oee.cell(row=row, column=CONES_COL[col_name]).value
            try:
                return float(v) if v is not None else default
            except (ValueError, TypeError):
                return default

        target_val = num("target_pcs")
        actual_val = num("actual_good_pcs")

        if target_val <= 0 or actual_val <= 0:
            continue

        machine_no = num("machine_no")
        if machine_no == 0:
            continue

        duration_hrs   = num("duration_hrs")
        run_hrs        = num("run_hrs")
        rated_speed    = num("rated_speed")
        weight_gm      = num("weight_gm")
        broken_gen_kg  = num("broken_gen_kg")
        broken_used_kg = num("broken_used_kg")

        # broken_gen_pcs computed from raw KG + weight (bypasses cached formula in col T)
        # Quality only uses broken GENERATED — broken_used (recycled) is excluded.
        broken_gen_pcs = (broken_gen_kg * 1000 / weight_gm) if weight_gm > 0 else 0.0

        # Actual speed = (good + broken_gen_pcs) / run_hrs  [for reference only]
        actual_speed = (actual_val + broken_gen_pcs) / run_hrs if run_hrs > 0 else 0.0

        records.append({
            "date":             date_str,
            "shift":            str(get("shift") or "").strip(),
            "machine_no":       int(machine_no),
            "operator_name":    str(get("operator_name") or "").strip(),
            "product_type":     str(get("product_type") or "").strip(),
            "cone_size":        str(get("cone_size") or "").strip(),
            "cone_type":        str(get("cone_type") or "").strip(),
            "duration_hrs":     duration_hrs,
            "run_hrs":          run_hrs,
            "rated_speed":      rated_speed,
            "actual_speed":     round(actual_speed, 2),
            "target_pcs":       target_val,
            "actual_good_pcs":  actual_val,
            "broken_gen_pcs":   round(broken_gen_pcs, 1),
            "broken_gen_kg":    broken_gen_kg,
            "broken_used_kg":   broken_used_kg,
            "dt_pump_clean":   num("dt_pump_clean"),
            "dt_mould_chg":    num("dt_mould_chg"),
            "dt_variety_co":   num("dt_variety_co"),
            "dt_gas_filter":   num("dt_gas_filter"),
            "dt_plate_clean":  num("dt_plate_clean"),
            "dt_proc_fail":    num("dt_proc_fail"),
            "dt_mach_bkdn":    num("dt_mach_bkdn"),
            "dt_manpower":     num("dt_manpower"),
            "dt_rm_pm":        num("dt_rm_pm"),
            "dt_power_cut":    num("dt_power_cut"),
            "dt_other":        num("dt_other"),
        })

    log.info(f"  OEE tracker:   {len(records)} active machine-shift rows")

    # Rebuild not_run to cover all slots not in OEE records
    ran_set = {(r["date"], r["shift"], r["machine_no"]) for r in records}
    reason_lookup = {
        (nr["date"], nr["shift"], nr["machine_no"]): nr["reason"]
        for nr in not_run
    }
    not_run = [
        {
            "date":       s["date"],
            "shift":      s["shift"],
            "machine_no": s["machine_no"],
            "reason":     reason_lookup.get(
                              (s["date"], s["shift"], s["machine_no"]),
                              "Missing reason code"
                          ),
        }
        for s in all_slots
        if (s["date"], s["shift"], s["machine_no"]) not in ran_set
    ]
    log.info(f"  Not-run slots: {len(not_run)}")

    return {
        "date_str":  date_str,
        "all_slots": all_slots,
        "not_run":   not_run,
        "records":   records,
    }


# ── Metrics computation ───────────────────────────────────────────────────────

def compute_metrics(records, not_run, all_slots):
    """
    3-factor OEE: Availability x Performance x Quality = OEE_running

    All metrics use aggregate sum-based ratios (not per-row weighted averages).

    OEE_running   = sum(good) / sum(target)
    Availability  = sum(run_hrs) / sum(shift_hrs)
    Quality       = sum(good) / sum(good + broken_gen_pcs)   [broken_gen only]
    Performance   = OEE_running / (Availability x Quality)   [back-calculated]
                    => guarantees Avail x Perf x Quality = OEE_running exactly
                    => negative perf_loss means above rated speed (allowed)
    Shift run %   = sum(hours_run) / (total_slots x 12)
    Overall OEE   = OEE_running x shift_run_pct
    """
    def s(fn):
        return sum(fn(r) for r in records)

    total_shift_hrs      = s(lambda r: r["duration_hrs"])
    total_run_hrs        = s(lambda r: r["run_hrs"])
    total_target         = s(lambda r: r["target_pcs"])
    total_actual         = s(lambda r: r["actual_good_pcs"])
    total_broken_gen_pcs = s(lambda r: r["broken_gen_pcs"])
    total_broken_gen     = s(lambda r: r["broken_gen_kg"])
    total_broken_used    = s(lambda r: r["broken_used_kg"])

    # OEE running = total good / total target
    oee_running = total_actual / total_target if total_target > 0 else 0.0

    # Availability = total run hrs / total shift hrs
    avail_rate = total_run_hrs / total_shift_hrs if total_shift_hrs > 0 else 1.0
    avail_loss = 1.0 - avail_rate

    # Quality = good / (good + broken_generated)  — broken_used excluded
    qual_denom   = total_actual + total_broken_gen_pcs
    quality_rate = total_actual / qual_denom if qual_denom > 0 else 1.0
    quality_loss = 1.0 - quality_rate

    # Performance = back-calculated residual so Avail x Perf x Quality = OEE exactly
    denom     = avail_rate * quality_rate
    perf_rate = oee_running / denom if denom > 0 else 1.0
    perf_loss = 1.0 - perf_rate  # negative = machines ran above rated speed (allowed)

    # Shift run %
    SHIFT_HRS     = 12.0
    total_slots   = len(all_slots)
    ran_keys_set  = {(r["date"], r["shift"], r["machine_no"]) for r in records}
    ran_cnt       = len(ran_keys_set)
    not_ran_cnt   = total_slots - ran_cnt
    total_ran_hrs = sum(s2.get("hours_run", SHIFT_HRS) for s2 in all_slots)
    shift_run_pct = total_ran_hrs / (total_slots * SHIFT_HRS) if total_slots else 1

    # Overall OEE = OEE_running x shift_run_pct
    overall_oee = oee_running * shift_run_pct

    # Downtime breakdown
    dt_breakdown_pct = {}
    dt_breakdown_hrs = {}
    for k, label in zip(DT_KEYS, DT_LABELS):
        hrs = s(lambda r, k=k: r[k])
        dt_breakdown_pct[label] = round(hrs / total_shift_hrs, 6) if total_shift_hrs else 0
        dt_breakdown_hrs[label] = round(hrs, 2)

    # Not-run reasons
    reason_counts = {}
    for nr in not_run:
        key = nr["reason"] or "Missing reason code"
        reason_counts[key] = reason_counts.get(key, 0) + 1
    reason_pcts = {
        k: round(v / total_slots, 6) if total_slots else 0
        for k, v in reason_counts.items()
    }

    return {
        "total_shift_hrs":      round(total_shift_hrs, 1),
        "total_run_hrs":        round(total_run_hrs, 1),
        "total_target":         int(total_target),
        "total_actual":         int(total_actual),
        "total_broken_gen_pcs": round(total_broken_gen_pcs, 0),
        "total_rej":            round(total_broken_gen_pcs, 0),
        "total_broken_gen":     round(total_broken_gen, 2),
        "total_broken_used":    round(total_broken_used, 2),
        "total_slots":          total_slots,
        "ran_slots":            ran_cnt,
        "not_ran_slots":        not_ran_cnt,
        "shift_run_pct":        round(shift_run_pct, 6),
        "avail_loss":           round(avail_loss, 6),
        "perf_loss":            round(perf_loss, 6),
        "quality_loss":         round(quality_loss, 6),
        "oee_running":          round(oee_running, 6),
        "overall_oee":          round(overall_oee, 6),
        "dt_breakdown_pct":     dt_breakdown_pct,
        "dt_breakdown_hrs":     dt_breakdown_hrs,
        "reason_pcts":          reason_pcts,
        "reason_counts":        reason_counts,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="IML OEE ETL -- converts daily Excel files to dashboard JSON"
    )
    parser.add_argument(
        "--folder",
        default=".",
        help='Folder containing daily Excel files (default: same folder as this script)'
    )
    parser.add_argument(
        "--output",
        default="oee_data.json",
        help="Output JSON path (default: oee_data.json)"
    )
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.exists():
        log.error(f"Folder not found: {folder}")
        sys.exit(1)

    excel_files = sorted(
        list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")),
        key=lambda p: extract_date(str(p)) or datetime.min
    )
    log.info(f"Found {len(excel_files)} Excel file(s) in {folder.resolve()}")

    if not excel_files:
        log.error("No .xlsx files found. Check the --folder path.")
        sys.exit(1)

    all_records  = []
    all_not_run  = []
    all_slots    = []
    daily_data   = []
    failed_files = []

    for fp in excel_files:
        result = parse_file(str(fp))
        if result is None:
            failed_files.append(fp.name)
            continue

        all_records.extend(result["records"])
        all_not_run.extend(result["not_run"])
        all_slots.extend(result["all_slots"])

        day_metrics = compute_metrics(
            result["records"], result["not_run"], result["all_slots"]
        )
        daily_data.append({
            "date":          result["date_str"],
            "overall_oee":   day_metrics["overall_oee"],
            "oee_running":   day_metrics["oee_running"],
            "shift_run_pct": day_metrics["shift_run_pct"],
            "avail_loss":    day_metrics["avail_loss"],
            "perf_loss":     day_metrics["perf_loss"],
            "quality_loss":  day_metrics["quality_loss"],
            "total_actual":  day_metrics["total_actual"],
            "total_target":  day_metrics["total_target"],
        })

    if not all_records:
        log.error("No valid data could be read from any file.")
        sys.exit(1)

    log.info(f"Computing aggregate across {len(daily_data)} date(s)...")
    aggregate = compute_metrics(all_records, all_not_run, all_slots)

    # Machine-level aggregation
    machine_nos = sorted(set(r["machine_no"] for r in all_records))
    machines = []
    for m in machine_nos:
        m_recs  = [r for r in all_records if r["machine_no"] == m]
        m_nr    = [r for r in all_not_run  if r["machine_no"] == m]
        m_slots = [r for r in all_slots    if r["machine_no"] == m]
        mc = compute_metrics(m_recs, m_nr, m_slots)

        sizes = [r["cone_size"] for r in m_recs if r["cone_size"]]
        primary_size = max(set(sizes), key=sizes.count) if sizes else ""
        types = [r["product_type"] for r in m_recs if r["product_type"]]
        primary_type = max(set(types), key=types.count) if types else ""
        ops = [r["operator_name"] for r in m_recs if r["operator_name"]]
        primary_op = max(set(ops), key=ops.count) if ops else ""

        machines.append({
            "machine_no":       m,
            "primary_size":     primary_size,
            "primary_type":     primary_type,
            "primary_operator": primary_op,
            "overall_oee":      mc["overall_oee"],
            "oee_running":      mc["oee_running"],
            "shift_run_pct":    mc["shift_run_pct"],
            "avail_loss":       mc["avail_loss"],
            "perf_loss":        mc["perf_loss"],
            "quality_loss":     mc["quality_loss"],
            "total_actual":     mc["total_actual"],
            "total_target":     mc["total_target"],
            "total_rej":        mc["total_rej"],
            "ran_slots":        mc["ran_slots"],
            "total_slots":      mc["total_slots"],
            "dt_breakdown_hrs": mc["dt_breakdown_hrs"],
        })

    date_range = {
        "from": daily_data[0]["date"] if daily_data else "",
        "to":   daily_data[-1]["date"] if daily_data else "",
        "days": len(daily_data),
    }

    output = {
        "_meta": {
            "generated":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_files": len(daily_data),
            "failed_files": failed_files,
            "date_range":   date_range,
            "formula_notes": {
                "overall_oee":  "OEE_running x shift_run_pct",
                "oee_running":  "sum(actual_good) / sum(target)",
                "shift_run_pct":"sum(hours_run) / (total_slots x 12)",
                "avail_loss":   "1 - sum(run_hrs) / sum(shift_hrs)",
                "quality_loss": "1 - sum(good) / sum(good + broken_gen_pcs)",
                "perf_loss":    "1 - oee_running / (avail_rate x quality_rate)  [back-calculated]",
                "downtime_pct": "each category hrs / total shift hrs",
            },
        },
        "date_range": date_range,
        "aggregate":  aggregate,
        "daily":      daily_data,
        "machines":   machines,
        "records":    all_records,
        "not_run":    all_not_run,
        "all_slots":  all_slots,
    }

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, default=str)

    log.info(f"\nDone. Written to: {out_path.resolve()}")
    log.info(f"  Dates:   {date_range['from']}  to  {date_range['to']}  ({date_range['days']} day(s))")
    log.info(f"  Records: {len(all_records)} machine-shift rows across {len(machines)} machines")

    if failed_files:
        log.warning(f"  Skipped {len(failed_files)} file(s): {failed_files}")

    # Console waterfall summary
    d = aggregate
    print("\n" + "="*58)
    print(f"  WATERFALL  ({date_range['from']}  to  {date_range['to']})")
    print("="*58)
    sr = d["shift_run_pct"]
    av = 1 - d["avail_loss"]
    pf = 1 - d["perf_loss"]
    ql = 1 - d["quality_loss"]
    print(f"  Theoretical max             100.00%")
    print(f"  x Shift run %          {sr*100:8.2f}%   loss {(1-sr)*100:.2f}%  ({d['not_ran_slots']}/{d['total_slots']} slots idle)")
    print(f"  x Availability         {sr*av*100:8.2f}%   loss {d['avail_loss']*100:.2f}%  (downtime)")
    print(f"  x Performance          {sr*av*pf*100:8.2f}%   loss {d['perf_loss']*100:.2f}%  (speed, back-calc)")
    print(f"  x Quality              {sr*av*pf*ql*100:8.2f}%   loss {d['quality_loss']*100:.2f}%  (rejections)")
    print(f"  = Overall OEE          {d['overall_oee']*100:8.2f}%")
    print("="*58)


if __name__ == "__main__":
    main()
